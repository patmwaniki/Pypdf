import PyPDF2
import os
import re
import openpyxl
import sys

path = 'C:\\Users\\Patrick.RAIPLY\Documents\\Work\\IT\\python\\mra_epz_f.xlsx'
your_excel =openpyxl.load_workbook(path)
your_sheet= your_excel['Sheet']



for file_name in os.listdir('all_format_pdf'): #Loop all file
	print(file_name)
	load_pdf = open(r'C:\\Users\\Patrick.RAIPLY\Documents\\Work\\IT\\python\\all_format_pdf\\'+file_name,'rb')
	read_pdf = PyPDF2.PdfFileReader(load_pdf)  #load 
	page_count = read_pdf.getNumPages()			
	first_page = read_pdf.getPage(0)			#Read first page only
	page_content = first_page.extractText()		#extra string output
	page_content = page_content.replace('\n','') #replace new line (if required)
	print(page_content)
	print('--------------------------------------------')

	#15 digit RC number 
	#Reg_1= '(?<= Receipt Number)'
	receipt_number = re.search(r'RC[A-Z\d]+', page_content).group()
	print(receipt_number)
	
	prn_date=  re.search(r'[A-Z\d]+', page_content).group()
	print(prn_date)

	try:
		prn_number = re.search(r'PRN Number : [\d \d\d\d\d \d\d\d\d \d\d\d\d]+', page_content).group()
	except:
		prn_number = re.search(r'PRN Number : [\d \d\d\d\d \d\d\d\d \d\d\d\d]+', page_content)
	
	print(prn_number)
	try:
		prn_total=  re.search(r'Grand Total [(.*?)\.+', page_content).group()
	except:
		prn_total=  re.search(r'Grand Total [(.*?)\.]+', page_content)
	print(prn_total)

	# last_row_number

	last_row_number = your_sheet.max_row
	#print(last_row_number)

	your_sheet.cell(column=1, row=last_row_number+1).value = prn_date
	your_sheet.cell(column=2, row=last_row_number+1).value = prn_number
	your_sheet.cell(column=3, row=last_row_number+1).value = receipt_number
	your_sheet.cell(column=4, row=last_row_number+1).value = prn_date
	your_sheet.cell(column=5, row=last_row_number+1).value = prn_total


	#save results the file 
	your_excel.save('C:\\Users\\Patrick.RAIPLY\Documents\\Work\\IT\\python\\mra_epz_f.xlsx')
	


	#doc_name = receipt_number + prn_number + prn_date + prn_total
	#print (doc_name)

	# last row 
	#last_row_number = your_sheet.max_row
	#print(last_row_number)
	#your_sheet = openpyxl.Workbook()



 #--
	#c1 = last_row_number.cell(row = 2, column = 1).value = prn_date
	#c2 = last_row_number.cell(row = 2, column = 2).value = prn_number
	#c3 = last_row_number.cell(row = 2, column = 3).value = receipt_number
	#c4 = last_row_number.cell(row = 2, column = 4).value = prn_date
	#c5 = last_row_number.cell(row = 2, column = 5).value = prn_total
	#for i in c1:
	#	last_row_number.cell(column=1,row=2+1)
	#your_sheet.save (path)

	#save results the file 
	#your_excel.save('C:\\Users\\Patrick.RAIPLY\Documents\\Work\\IT\\python\\mra_2.xlsx'
with open('readme.txt', 'w') as f:
    f.write('readme')
	